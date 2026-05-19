from __future__ import annotations

import concurrent.futures
import time
from datetime import datetime, timedelta, timezone
from typing import Optional

from fastapi import APIRouter, HTTPException, Query, Response
from fastapi.responses import JSONResponse

import state
from core import (
    _hik_call, _extract_records, _resolve_person_query, decrypt_in_place, get_client,
)

router = APIRouter()

_PAGE_SIZE = 500
_MAX_WORKERS = 6


def _fetch_all_pages(
    start_time: Optional[str],
    end_time: Optional[str],
    eids: str,
    pid: Optional[int] = None,
    pname: Optional[str] = None,
) -> tuple[list, int]:
    """Страница 1 — узнать total, остальные страницы — параллельно."""

    def _fetch(pg: int) -> dict:
        return _hik_call(
            lambda c, __p=pg: c.card_swipe_records(
                page=__p, page_size=_PAGE_SIZE,
                start_time=start_time, end_time=end_time,
                person_id=pid, person_name=pname,
                element_ids=eids,
            )
        )

    raw1 = _fetch(1)
    if raw1.get("ResponseStatus", {}).get("ErrorCode", 0) != 0:
        return [], 0

    batch1, total = _extract_records(raw1)
    if not batch1 or len(batch1) >= total:
        return batch1, total

    num_pages = (total + _PAGE_SIZE - 1) // _PAGE_SIZE
    remaining = list(range(2, num_pages + 1))
    page_results: dict[int, list] = {1: batch1}

    with concurrent.futures.ThreadPoolExecutor(max_workers=min(_MAX_WORKERS, len(remaining))) as ex:
        futures = {ex.submit(_fetch, pg): pg for pg in remaining}
        for fut in concurrent.futures.as_completed(futures):
            pg = futures[fut]
            try:
                batch, _ = _extract_records(fut.result())
                page_results[pg] = batch
            except Exception:
                page_results[pg] = []

    all_records: list = []
    for pg in range(1, num_pages + 1):
        all_records.extend(page_results.get(pg, []))

    return all_records, total


def _record_dedup_key(r: dict) -> str:
    t   = r.get("DeviceTime") or r.get("EventTime") or ""
    eid = str(r.get("ElementID") or r.get("DoorID") or "")
    pid = str((r.get("Person") or {}).get("ID") or "")
    return f"{t}|{eid}|{pid}"


def _hik_data(raw: dict) -> dict:
    import json as _json
    rs = raw.get("ResponseStatus", raw)
    data = rs.get("Data", rs)
    if isinstance(data, str):
        try:
            data = _json.loads(data)
        except Exception:
            data = {}
    return data if isinstance(data, dict) else {}


@router.get("/api/records", tags=["Records"])
def list_records(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=1000),
    fetch_all: bool = Query(False, description="Загрузить ВСЕ записи (бэкенд сам обходит все страницы)"),
    start_time: Optional[str] = Query(None, description="ISO 8601, e.g. 2026-05-14T00:00:00+05:00"),
    end_time: Optional[str] = Query(None),
    person_id: Optional[int] = Query(None),
    person_name: Optional[str] = Query(None),
    element_ids: Optional[str] = Query(None, description="ID точек доступа через запятую"),
):
    """Записи проходов (события доступа). fetch_all=true — вернуть весь список без пагинации."""
    _eids = element_ids or ""
    if not start_time and not end_time:
        _now = datetime.now(timezone.utc).astimezone()
        start_time = _now.replace(hour=0, minute=0, second=0, microsecond=0).isoformat(timespec="seconds")
        end_time = _now.isoformat(timespec="seconds")

    _pid, _multi_pids, _pname = _resolve_person_query(person_name, person_id)

    if fetch_all:
        try:
            if _multi_pids:
                # несколько персон — параллельно по каждому pid
                all_records: list = []
                _total = 0
                seen_ids: set = set()
                with concurrent.futures.ThreadPoolExecutor(max_workers=min(_MAX_WORKERS, len(_multi_pids))) as ex:
                    futures = {
                        ex.submit(_fetch_all_pages, start_time, end_time, _eids, pid, None): pid
                        for pid in _multi_pids
                    }
                    for fut in concurrent.futures.as_completed(futures):
                        _batch, _t = fut.result()
                        _total += _t
                        for r in _batch:
                            _key = _record_dedup_key(r)
                            if _key not in seen_ids:
                                seen_ids.add(_key)
                                all_records.append(r)
            else:
                all_records, _total = _fetch_all_pages(start_time, end_time, _eids, _pid, _pname)
        except Exception as e:
            raise HTTPException(500, str(e))

        decrypt_in_place(all_records, get_client().aes_key_hex)
        return {"page": 1, "page_size": len(all_records), "total": _total, "records": all_records}

    _paginated_pid = _multi_pids[0] if _multi_pids else _pid
    _paginated_pname = None if _multi_pids else _pname
    try:
        raw = _hik_call(lambda c: c.card_swipe_records(
            page=page, page_size=page_size,
            start_time=start_time, end_time=end_time,
            person_id=_paginated_pid, person_name=_paginated_pname,
            element_ids=_eids,
        ))
    except Exception as e:
        raise HTTPException(500, str(e))

    rs = raw.get("ResponseStatus", {})
    if rs and rs.get("ErrorCode", 0) != 0:
        return JSONResponse(status_code=400, content={"error": rs, "raw": raw})

    records, total = _extract_records(raw)
    decrypt_in_place(records, get_client().aes_key_hex)
    return {"page": page, "page_size": page_size, "total": total, "records": records}


@router.get("/api/records/export.xlsx", tags=["Records"])
def export_records_xlsx(
    start_time: Optional[str] = Query(None),
    end_time: Optional[str] = Query(None),
    person_id: Optional[int] = Query(None),
    person_name: Optional[str] = Query(None),
    element_ids: Optional[str] = Query(None),
):
    """Выгрузить проходы в Excel (.xlsx). Параметры те же что у /api/records."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    if not start_time and not end_time:
        _now = datetime.now(timezone.utc).astimezone()
        start_time = _now.replace(hour=0, minute=0, second=0, microsecond=0).isoformat(timespec="seconds")
        end_time = _now.isoformat(timespec="seconds")

    _pid, _multi_pids_export, _pname = _resolve_person_query(person_name, person_id)
    if _multi_pids_export:
        _pid = _multi_pids_export[0]
        _pname = None

    _eids = element_ids or ""
    try:
        all_records, _ = _fetch_all_pages(start_time, end_time, _eids, _pid, _pname)
    except Exception as e:
        raise HTTPException(500, str(e))

    decrypt_in_place(all_records, get_client().aes_key_hex)

    wb = Workbook()
    ws = wb.active
    ws.title = "Проходы"

    header_fill = PatternFill("solid", fgColor="1E2235")
    header_font = Font(bold=True, color="C8CCDE", size=11)
    headers = ["#", "Имя", "ИИН", "Код", "Отдел / Класс", "Время", "Точка доступа", "Считыватель", "Результат"]
    col_widths = [5, 32, 16, 12, 30, 22, 22, 20, 12]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = w

    ws.freeze_panes = "A2"
    auth_map = {1: "Разрешён", 0: "Отказ"}

    for ri, r in enumerate(all_records, 1):
        b = (r.get("Person") or {}).get("BaseInfo") or {}
        given = b.get("GivenName") or b.get("FullName") or ""
        iin   = b.get("FamilyName") or ""
        code  = b.get("PersonCode") or ""
        dept  = b.get("FullPath") or ""
        pid   = (r.get("Person") or {}).get("ID")
        name  = given or (f"ID {pid}" if pid else "—")
        dt_raw = r.get("DeviceTime") or ""
        try:
            dt_str = datetime.fromisoformat(dt_raw).strftime("%d.%m.%Y %H:%M:%S") if dt_raw else "—"
        except Exception:
            dt_str = dt_raw
        result = auth_map.get(r.get("SwipeAuthResult"), "—")
        row_data = [
            ri, name, iin, code, dept, dt_str,
            r.get("ElementName") or "—",
            r.get("CardReaderName") or "—",
            result,
        ]
        for ci, val in enumerate(row_data, 1):
            ws.cell(row=ri + 1, column=ci, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"records_{(start_time or '')[:10]}_{(end_time or '')[:10]}.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/api/access-points", tags=["Records"])
def list_access_points():
    """
    Список точек доступа.
    Приоритет: Elements API (с online-статусом). Fallback — сканирование записей за 30 дней.
    """
    def _elements_valid() -> bool:
        return (
            state._elements_cache["data"] is not None
            and time.time() - state._elements_cache["ts"] < state._ELEMENTS_CACHE_TTL_S
        )

    if _elements_valid():
        return {"access_points": state._elements_cache["data"]}

    with state._elements_lock:
        if _elements_valid():
            return {"access_points": state._elements_cache["data"]}

        try:
            raw = _hik_call(lambda c: c.logical_elements())
            el_list = _hik_data(raw).get("ElementList", {}).get("Element", [])
            if el_list:
                points = [
                    {
                        "id": str(e.get("ID", "")),
                        "guid": e.get("GUID", ""),
                        "name": e.get("Name", ""),
                        "type": e.get("Type"),
                        "online": bool(e.get("Online")),
                    }
                    for e in el_list
                ]
                points.sort(key=lambda x: x["name"])
                state._elements_cache["data"] = points
                state._elements_cache["ts"] = time.time()
                return {"access_points": points}
        except Exception:
            pass

    now = datetime.now(timezone.utc).astimezone()
    start = (now - timedelta(days=30)).replace(hour=0, minute=0, second=0, microsecond=0)
    seen: dict[str, dict] = {}
    _p, _ps = 1, 500
    try:
        while True:
            raw = _hik_call(lambda c, __p=_p: c.card_swipe_records(
                page=__p, page_size=_ps,
                start_time=start.isoformat(timespec="seconds"),
                end_time=now.isoformat(timespec="seconds"),
            ))
            batch, total = _extract_records(raw)
            if not batch:
                break
            for rec in batch:
                name = str(rec.get("ElementName") or "")
                eid = str(rec.get("ElementID") or rec.get("DoorID") or "")
                key = eid if eid else name
                if name and key and key not in seen:
                    seen[key] = {"id": eid if eid else name, "name": name, "online": None}
            if len(seen) >= total or len(batch) < _ps:
                break
            _p += 1
    except Exception as e:
        return {"access_points": [], "error": str(e)}

    return {"access_points": sorted(seen.values(), key=lambda x: x["name"])}
